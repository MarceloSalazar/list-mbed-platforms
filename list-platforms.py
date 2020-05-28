# Simple script to read target data from targets.json and spreadsheet

import requests
import os
import urllib, json
import time
import openpyxl.styles
from argparse import ArgumentParser
from prettytable import PrettyTable
from natsort import natsorted
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW, RED
from progress.bar import IncrementalBar

# Spradsheet cell location
XLS_COL_TARGET = 1
XLS_COL_GITHUB = 2
XLS_COL_GITHUB_RTOS = 3
XLS_COL_GITHUB_BARE = 4
XLS_COL_MBEDCOM_60  = 5
XLS_COL_MBEDCOM_RTOS = 6
XLS_COL_MBEDCOM_BARE = 7
XLS_COL_ME_BASELINE = 8
XLS_COL_ME_ADVANCED = 9
XLS_COL_ME_PELION = 10

os_mbed_com_data = {}

class Target_database(object):

    class Target(object):
          def __init__(self, target_name, github=False, \
                        mbedcom60=False, mbedcom6_rtos=False, mbedcom6_bare=False,
                        me_baseline = False, me_advanced= False, me_pelion=False):
              
            self.target_name = str(target_name).upper()
            self.board_name = ''
            self.mbed_com_url = ''
            self.target_github         = github  # target is on github or not
            self.target_github_rtos    = False   # target on github has rtos enabled or not (supported_application_profile)
            self.target_github_bare    = False   # target on github has baremetal enabled or not (supported_application_profile)
            self.target_mbed_com60     = mbedcom60 # target is public or not
            self.target_mbed_com6_rtos = mbedcom6_rtos # target has rtos feature or not
            self.target_mbed_com6_bare = mbedcom6_bare # target has baremetal feature or not
            self.target_me_baseline    = me_baseline
            self.target_me_advanced    = me_advanced
            self.target_me_pelion      = me_pelion
            
    TARGETS_IGNORE = ["TARGET", "PSA_TARGET", "NSPE_TARGET", "SPE_TARGET", "CM4_UARM", "CM4_ARM", \
                  "CM4F_UARM", "CM4F_ARM", "__BUILD_TOOLS_METADATA__"]
            
    def __init__(self):
        ''' Initializes database '''
        self.targets = [] # List of objects
                    
        # 1st read from Github
        self.update_from_github()
        
        # 2nd read from mbed.com and ignore targets that are not found
        self.update_from_mbed_com()

    def get_index(self, target):
        
        target = str(target).upper()
        
        for i in range(len(self.targets)):
            if target == self.targets[i].target_name:
                return i
        return False
        
    def add_target(self, target):
        ''' TODO '''

        if target.target_name in self.TARGETS_IGNORE:
            return
        
        # Check if item is in list
        target_not_found = True
        
        for i in range(len(self.targets)):
            if self.targets[i].target_name == target.target_name:
                target_not_found = False
                
                if target.target_github:
                    self.targets[i].target_github = target.target_github
                
                if target.target_mbed_com60:
                    self.targets[i].target_mbed_com60 = target.target_mbed_com60

                if target.target_mbed_com6_rtos:
                    self.targets[i].target_mbed_com6_rtos = target.target_mbed_com6_rtos

                if target.target_mbed_com6_bare:
                    self.targets[i].target_mbed_com6_bare = target.target_mbed_com6_bare                

                if target.target_me_baseline:
                    self.targets[i].target_me_baseline = target.target_me_baseline               
              
                if target.target_me_advanced:
                    self.targets[i].target_me_advanced = target.target_me_advanced    

                if target.target_me_pelion:
                    self.targets[i].target_me_pelion = target.target_me_pelion 
                    
                         
        if target_not_found:
            self.targets.append(target)   
    
    
    def update_from_github(self, branch='master'):
        ''' TODO '''
        
        # TODO: check branch/tag
        url_targets_json = 'https://raw.github.com/ARMmbed/mbed-os/master/targets/targets.json'

        res = requests.get(url_targets_json)
        db = res.json()

        print("\nRead from github...\n")
        bar = IncrementalBar('Processing', max=len(db))

        # TODO
        # - Check if target is RTOS
        # - Check if target is Baremetal

        for i in db:
            bar.next()
            if 'public' in db[i] and db[i]['public'] == False:
                continue
            new_target = self.Target(i, github=True)
            self.add_target(new_target)
     
    def update_from_mbed_com(self):

        targets = {}
        # Read data from os.mbed.com and crete local database
        # TODO: use user/pass using dotenv info
        
        url_os_mbed_com = 'https://os.mbed.com/api/v3/platforms/'
        response = requests.get(url_os_mbed_com,auth=('user', 'password'))
        db = response.json()
                
        print("\nRead from mbed.com...\n")
        bar = IncrementalBar('Processing', max=len(db))
        for i in range(len(db)):
            bar.next()
            
            m60      = False
            m6RTOS   = False
            m6BARE   = False
            meBase   = False
            meAdv    = False
            mePelion = False

            for j in db[i]['features']:
                if j['category']['name'] == "Mbed OS support":                    
                    if j['name'] == "Mbed OS 5.15":
                        m60 = True

                if j['category']['name'] == "Mbed OS 6":                    
                    if j['name'] == "RTOS":
                        m6RTOS = True
                    if j['name'] == "Bare metal":
                        m6BARE = True

                if j['category']['name'] == "Mbed Enabled":                    
                    if j['name'] == "Baseline":
                        meBase = True
                    if j['name'] == "Advanced":
                        meAdv  = True
                    if j['name'] == "Pelion Device Ready":
                        mePelion = True
                        
            new_target = self.Target(db[i]['logicalboard']['name'], \
                mbedcom60=m60, mbedcom6_rtos=m6RTOS , mbedcom6_bare=m6BARE, \
                me_baseline=meBase, me_advanced=meAdv, me_pelion=mePelion )

            
            # TODO:
            # - check if it's an Mbed OS 6 flag ==> all should be detected as false until public release
            # - check if Mbed 6 RTOS is set
            # - check if Mbed 6 Baremetal is set
            # - check if Mbed Enabled (baseline, advanced, pelion) is set
            
            self.add_target(new_target)

            
        return
            
    def print_table(self):

        table_header = ['#', 'Target name', \
                             'GitHub', 'GitHub RTOS', 'GitHub Bare', \
                             'Mbed.com 6.0', 'Mbed.com RTOS', 'Mbed.com Bare', \
                             'ME Baseline',  'ME Advanced', 'ME Pelion-Rdy']
        
        table = PrettyTable(table_header)
        table.align['Target name'] = 'l'
        for i in range(len(self.targets)):
            if self.targets[i].target_github:
                table.add_row([i, \
                               self.targets[i].target_name, \
                               self.targets[i].target_github, '', '', \
                               self.targets[i].target_mbed_com60, '', '',\
                               self.targets[i].target_me_baseline, \
                               self.targets[i].target_me_advanced, \
                               self.targets[i].target_me_pelion])
        print("\n")
        print(table)

    def update_spreadsheet(self, spreadsheet_file):

        # TODO
        # Check all flags: Github, RTOS, Baremetal

        if spreadsheet_file == '':
            print("Spreadsheet not found")
            return

        # Open spreadsheet
        try:
            wb = load_workbook(filename = spreadsheet_file)
            ws = wb.active
        except:
            print("Could not find '.xlsx' in the current directory. Skipping check.")
            return
    
        # 1st read targets from spreadsheet until cell is blank
        xls_targets = []
        for i in range(2, ws.max_row+1):
            read_target = str(ws.cell(row=i, column=XLS_COL_TARGET).value).upper()
            if read_target == 'NONE' or read_target == '':
                break
            xls_targets.append(read_target)
        
        # 2nd add rest of targets into spreadsheet
        row_write = ws.max_row+1
        for i in range(len(self.targets)):
            if self.targets[i].target_name not in xls_targets:
                if self.targets[i].target_github:
                    ws.cell(row=row_write, column=XLS_COL_TARGET).value = self.targets[i].target_name
                    row_write += 1

        wb.save(filename = spreadsheet_file)    

        # 3rd update all fields for all spreadsheet
        fill_red = openpyxl.styles.fills.PatternFill(fgColor=RED, fill_type = "solid")
        fill_yel = openpyxl.styles.fills.PatternFill(fgColor=YELLOW, fill_type = "solid")
        fill_non = openpyxl.styles.fills.PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
        
        # clear background for all cells
        for i in range(2, ws.max_row+1):
            for j in range(1,20):
                ws.cell(row=i, column=j).fill = fill_non

        print("\nRead targets from spreadsheet")
        bar = IncrementalBar('Processing', max=ws.max_row)
        
        for i in range(2, ws.max_row+1):
            read_target = str(ws.cell(row=i, column=XLS_COL_TARGET).value).upper()
            idx = self.get_index(read_target)
            
            # Github
            ws.cell(row=i, column=XLS_COL_GITHUB).value = self.targets[idx].target_github            
            
            # Github RTOS
            # TODO
            
            # Github Baremetal
            # TODO
            
            # Mbed.com 6.0
            ws.cell(row=i, column=XLS_COL_MBEDCOM_60).value = self.targets[idx].target_mbed_com60  
            if self.targets[idx].target_mbed_com60 == False:
                ws.cell(row=i, column=XLS_COL_MBEDCOM_60).fill = fill_red

            # Mbed.com 6 RTOS
            ws.cell(row=i, column=XLS_COL_MBEDCOM_RTOS).value = self.targets[idx].target_mbed_com6_rtos

            # Mbed.com 6 Bare
            ws.cell(row=i, column=XLS_COL_MBEDCOM_BARE).value = self.targets[idx].target_mbed_com6_bare

            # Mbed Enabled Baseline
            ws.cell(row=i, column=XLS_COL_ME_BASELINE).value = self.targets[idx].target_me_baseline
            if self.targets[idx].target_me_advanced or self.targets[idx].target_me_pelion:
                if self.targets[idx].target_me_baseline == False:
                    ws.cell(row=i, column=XLS_COL_ME_BASELINE).fill = fill_yel
                    
            # Mbed Enabled Advanced
            ws.cell(row=i, column=XLS_COL_ME_ADVANCED).value = self.targets[idx].target_me_advanced
            if self.targets[idx].target_me_pelion:
                if self.targets[idx].target_me_advanced == False:
                    ws.cell(row=i, column=XLS_COL_ME_ADVANCED).fill = fill_yel
                    
            # Mbed Enabled Pelion
            ws.cell(row=i, column=XLS_COL_ME_PELION).value = self.targets[idx].target_me_pelion
            
            bar.next()        
        
        wb.save(filename = spreadsheet_file)   

def main():
    global args

    # Parser handling
    parser = ArgumentParser(description="Data parser from os.mbed.com/platforms")

    parser.add_argument(
        '-f', '--file', dest='file',
        help='XLSX file to update with all information', required=False)

    args = parser.parse_args()

    db = Target_database()

    # Load spreadsheet and update
    if args.file:
         db.update_spreadsheet(args.file)
    else:
        db.print_table()

    exit(0)

if __name__ == "__main__":
    main()